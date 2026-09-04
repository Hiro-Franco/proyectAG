import os, json, base64, datetime, re, threading
from functools import wraps
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import bcrypt, jwt, pytesseract
from PIL import Image
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app, origins='*')

SECRET_KEY  = os.environ.get('SECRET_KEY', 'cambia-esta-clave')
USERS_FILE  = os.path.join(os.path.dirname(__file__), '..', 'json', 'users.json')

# Archivo Excel donde se van acumulando todas las facturas.
# Podés apuntarlo a una carpeta compartida/sincronizada (Drive, OneDrive, etc.)
# seteando la variable de entorno EXCEL_FILE.
EXCEL_FILE = os.environ.get(
    'EXCEL_FILE',
    os.path.join(os.path.dirname(__file__), '..', 'data', 'facturas.xlsx')
)

EXCEL_HEADERS = [
    'Timestamp', 'Usuario', 'Fecha emisión', 'Razón social', 'Nombre comercial',
    'Categoría', 'Método de pago', 'Banco', 'CBU', 'Notas',
    'N° Factura', 'Fecha factura', 'CUIT', 'Total', 'Descripción',
]

# Varias facturas pueden llegar casi al mismo tiempo (varios operadores
# usando la app a la vez). Este lock evita que dos escrituras al mismo
# .xlsx se pisen entre sí y corrompan el archivo.
excel_lock = threading.Lock()


def parsear_importe(texto):
    """Convierte '$15.000', '15.000,50' o '' a un float. Formato AR
    (punto = miles, coma = decimales)."""
    if not texto:
        return 0.0
    limpio = str(texto).replace('$', '').strip()
    try:
        return float(limpio.replace('.', '').replace(',', '.'))
    except ValueError:
        return 0.0


def escribir_en_excel(fila):
    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
    with excel_lock:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        # --- Hoja "Facturas": historial completo, orden cronológico ---
        if 'Facturas' not in wb.sheetnames:
            ws = wb.create_sheet('Facturas')
            ws.append(EXCEL_HEADERS)
            for i, _ in enumerate(EXCEL_HEADERS, start=1):
                ws.cell(row=1, column=i).font = Font(bold=True)
            ws.freeze_panes = 'A2'
        else:
            ws = wb['Facturas']
        ws.append(fila)

        # --- Hoja "Resumen": un total por categoría, se recalcula sola ---
        actualizar_resumen(wb, ws)

        wb.save(EXCEL_FILE)


def actualizar_resumen(wb, ws_facturas):
    idx_categoria = EXCEL_HEADERS.index('Categoría')
    idx_total     = EXCEL_HEADERS.index('Total')

    totales = {}
    cantidades = {}
    for fila in ws_facturas.iter_rows(min_row=2, values_only=True):
        categoria = (fila[idx_categoria] or 'Sin categoría').strip() or 'Sin categoría'
        totales[categoria]    = totales.get(categoria, 0.0) + parsear_importe(fila[idx_total])
        cantidades[categoria] = cantidades.get(categoria, 0) + 1

    if 'Resumen' in wb.sheetnames:
        wb.remove(wb['Resumen'])
    ws = wb.create_sheet('Resumen', 0)  # la dejamos primera, es la que se mira más

    ws.append(['Categoría', 'Cantidad de facturas', 'Total'])
    for c in range(1, 4):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for categoria in sorted(totales.keys()):
        ws.append([categoria, cantidades[categoria], totales[categoria]])

    fila_total = len(totales) + 2
    ws.cell(row=fila_total, column=1, value='TOTAL GENERAL').font = Font(bold=True)
    ws.cell(row=fila_total, column=2, value=sum(cantidades.values())).font = Font(bold=True)
    ws.cell(row=fila_total, column=3, value=sum(totales.values())).font = Font(bold=True)

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '$ #,##0.00'

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.freeze_panes = 'A2'

def cargar_usuarios():
    with open(USERS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)['users']

def generar_token(username):
    payload = {
        'sub': username,
        'iat': datetime.datetime.utcnow(),
        'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=8),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm='HS256')

def verificar_token(token):
    try:    return jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
    except: return None

def requiere_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        if not auth.startswith('Bearer '): return jsonify({'error': 'Token requerido'}), 401
        payload = verificar_token(auth.split(' ', 1)[1])
        if not payload: return jsonify({'error': 'Token inválido o expirado'}), 401
        request.usuario = payload['sub']
        return f(*args, **kwargs)
    return decorated

@app.route('/auth/login', methods=['POST'])
def login():
    data     = request.get_json()
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').encode('utf-8')
    usuarios = cargar_usuarios()
    usuario  = next((u for u in usuarios if u['username'] == username), None)
    if not usuario or not bcrypt.checkpw(password, usuario['password_hash'].encode('utf-8')):
        return jsonify({'error': 'Usuario o contraseña incorrectos'}), 401
    return jsonify({'token': generar_token(username), 'nombre': usuario.get('nombre', username)}), 200

@app.route('/ocr/procesar', methods=['POST'])
@requiere_auth
def ocr_procesar():
    data = request.get_json()
    b64  = data.get('imagen', '')
    if not b64: return jsonify({'error': 'No se recibió imagen'}), 400
    try:
        imagen  = Image.open(io.BytesIO(base64.b64decode(b64))).convert('L')
        texto   = pytesseract.image_to_string(imagen, lang='spa', config='--oem 3 --psm 6')
        return jsonify(extraer_campos(texto)), 200
    except Exception as e:
        return jsonify({'error': f'Error OCR: {str(e)}'}), 500

def extraer_campos(texto):
    campos = {'numero': '', 'fecha_factura': '', 'cuit': '', 'total': '', 'descripcion': ''}
    m = re.search(r'(?:N[°º]?\s*|Factura\s+)[:\s]*(\d{4}[-\s]\d{8})', texto, re.I)
    if m: campos['numero'] = m.group(1).strip()
    m = re.search(r'(?:Fecha|Date)\s*[:\s]*(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})', texto, re.I)
    if m: campos['fecha_factura'] = m.group(1).strip()
    m = re.search(r'CUIT\s*[:\s]*(\d{2}[-\s]\d{8}[-\s]\d)', texto, re.I)
    if m: campos['cuit'] = m.group(1).replace(' ', '-').strip()
    importes = re.findall(r'\$\s*([\d.,]+)', texto)
    if importes:
        def p(s):
            try: return float(s.replace('.','').replace(',','.'))
            except: return 0
        campos['total'] = '$' + max(importes, key=p)
    for linea in [l.strip() for l in texto.split('\n') if l.strip()]:
        if not re.match(r'^[\d$%\-/]+$', linea) and len(linea) > 4:
            campos['descripcion'] = linea[:120]; break
    return campos

@app.route('/excel/guardar', methods=['POST'])
@requiere_auth
def excel_guardar():
    data = request.get_json()
    try:
        escribir_en_excel([
            data.get('timestamp',''), data.get('usuario',''),
            data.get('fecha',''), data.get('razonSocial',''), data.get('nombreComercial',''),
            data.get('categoria',''), data.get('metodoPago',''), data.get('banco',''),
            data.get('cbu',''), data.get('notas',''),
            data.get('numero',''), data.get('fecha_factura',''),
            data.get('cuit',''), data.get('total',''), data.get('descripcion',''),
        ])
        return jsonify({'ok': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/excel/descargar', methods=['GET'])
@requiere_auth
def excel_descargar():
    if not os.path.exists(EXCEL_FILE):
        return jsonify({'error': 'Todavía no hay facturas guardadas'}), 404
    return send_file(EXCEL_FILE, as_attachment=True, download_name='facturas.xlsx')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)